import openpyxl  # Importa a biblioteca openpyxl, usada para criar e manipular arquivos Excel (.xlsx)

# Criar planilha
# Sheet são páginas na planilha
planilha_test = openpyxl.Workbook()  # Cria uma nova planilha do Excel (workbook) vazia
planilha_test.create_sheet('Frutas')  # Adiciona uma nova aba chamada 'Frutas' ao workbook
planilha_test.create_sheet('Legumes')  # Adiciona uma nova aba chamada 'Legumes' ao workbook
planilha_test.create_sheet('Sementes')  # Adiciona uma nova aba chamada 'Sementes' ao workbook
print(planilha_test.sheetnames)  # Exibe uma lista com os nomes de todas as abas criadas no workbook

# Adicionar Cabeçalho
cabecalho_frutas = ['Título', 'Localização', 'Preço']  # Define uma lista com os títulos das colunas (cabeçalho)
sheet_frutas = planilha_test['Frutas']  # Obtém a aba 'Frutas' do workbook
sheet_frutas.append(cabecalho_frutas)  # Adiciona o cabeçalho como a primeira linha na aba 'Frutas'

##################################################

frutas = [['Uva', 'Mercado', 5],['Melancia', 'Mercado', 15],['Bolo', 'Mercado', 40]]

for fruta in frutas:
    sheet_frutas.append(fruta)

# Salvando planilha
planilha_test.save('Ex21_Planilhas_01/Dados_Supermercado.xlsx')  # Salva o workbook no arquivo especificado, criando o arquivo Excel
